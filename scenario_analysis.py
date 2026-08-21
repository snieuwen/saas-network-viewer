from __future__ import annotations

import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import tkinter as tk
import tkinter.font as tkfont
from tkinter import messagebox, ttk

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from user_analysis import UserAssignmentDialog, scenario_exclusion_mask, scenario_user_summary


SCENARIO_CATEGORY_ORDER = {
    "Manage": 0,
    "Inkoop Raadplegen": 1,
    "Contract Raadplegen": 2,
    "Raadplegen": 3,
    "Projecten": 4,
}
SCENARIO_SHEETS = ("Roles", "Privileges", "Privilege-Roles")
DEFAULT_SCENARIO_COMMENT = (
    "# Rows whose first cell starts with # are comments/group headings; ignored."
)


@dataclass
class Scenario:
    name: str
    sku: str
    service: str
    excluded_roles: list[str] = field(default_factory=list)
    excluded_privileges: list[str] = field(default_factory=list)
    excluded_privilege_roles: list[dict[str, str]] = field(default_factory=list)
    sheet_comments: dict[str, str] = field(default_factory=dict)
    created_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))

    def add_nodes(self, node_ids: list[str]) -> None:
        for node_id in node_ids:
            kind, label = node_id.split(":", 1)
            target = self.excluded_roles if kind == "role" else self.excluded_privileges
            if label not in target:
                target.append(label)
        self.excluded_roles.sort()
        self.excluded_privileges.sort()

    def add_privilege_roles(self, privilege: str, roles: list[str], category: str = "") -> None:
        existing = {
            (item.get("privilege", ""), item.get("role", ""))
            for item in self.excluded_privilege_roles
        }
        for role in roles:
            pair = (privilege, role)
            if pair not in existing:
                item = {"privilege": privilege, "role": role}
                if category:
                    item["category"] = category
                self.excluded_privilege_roles.append(item)
                existing.add(pair)
        self.excluded_privilege_roles.sort(
            key=lambda item: (
                SCENARIO_CATEGORY_ORDER.get(item.get("category", ""), 999),
                item.get("category", ""),
                item.get("privilege", ""),
                item.get("role", ""),
            )
        )


def scenario_impact(frame: pd.DataFrame, scenario: Scenario | None = None) -> dict[str, int]:
    """Return the access-based user impact after scenario exclusions."""
    baseline_users = set(frame["USER_LOGIN_HASH"].astype(str)) if not frame.empty else set()
    if scenario is None:
        return {"in_scope": len(baseline_users), "removed": 0, "affected": 0}
    excluded_mask = scenario_exclusion_mask(frame, scenario)
    excluded = frame[excluded_mask]
    remaining = frame.drop(excluded.index)
    remaining_users = set(remaining["USER_LOGIN_HASH"].astype(str))
    return {
        "in_scope": len(remaining_users),
        "removed": len(baseline_users - remaining_users),
        "affected": int(excluded["USER_LOGIN_HASH"].nunique()),
    }


class ScenarioLibrary:
    def __init__(self) -> None:
        self.scenarios: list[Scenario] = []
        application_dir = (
            Path(sys.executable).resolve().parent
            if getattr(sys, "frozen", False)
            else Path(__file__).resolve().parent
        )
        self.storage_dir = application_dir / "scenarios"

    def for_sku(self, sku: str, service: str) -> list[Scenario]:
        return [item for item in self.scenarios if item.sku == sku and item.service == service]

    @staticmethod
    def _filename(name: str) -> str:
        cleaned = "".join("_" if char in '<>:"/\\|?*' else char for char in name).strip(" .")
        return f"{cleaned or 'Untitled scenario'}.xlsx"

    def save_scenario(self, scenario: Scenario) -> None:
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        path = self.storage_dir / self._filename(scenario.name)
        workbook = Workbook()
        roles_sheet = workbook.active
        roles_sheet.title = "Roles"
        privileges_sheet = workbook.create_sheet("Privileges")
        pairs_sheet = workbook.create_sheet("Privilege-Roles")
        self._write_scenario_sheet(
            roles_sheet,
            ["Role"],
            [[value] for value in scenario.excluded_roles],
            scenario.sheet_comments.get("Roles", ""),
        )
        metadata = (
            ("Scenario", scenario.name),
            ("SKU", scenario.sku),
            ("Service", scenario.service),
            ("Created at", scenario.created_at),
        )
        for row, (label, value) in enumerate(metadata, start=1):
            roles_sheet.cell(row, 4, label)
            roles_sheet.cell(row, 5, value)
            roles_sheet.cell(row, 4).font = Font(bold=True, color="59636E")
            roles_sheet.cell(row, 4).alignment = Alignment(horizontal="left")
            roles_sheet.cell(row, 5).alignment = Alignment(horizontal="left")
        roles_sheet.column_dimensions["B"].width = 2
        roles_sheet.column_dimensions["C"].width = 2
        roles_sheet.column_dimensions["D"].width = 16
        roles_sheet.column_dimensions["E"].width = 78
        self._write_scenario_sheet(
            privileges_sheet,
            ["Privilege"],
            [[value] for value in scenario.excluded_privileges],
            scenario.sheet_comments.get("Privileges", ""),
        )
        self._write_scenario_sheet(
            pairs_sheet,
            ["Privilege", "Role"],
            self._privilege_role_rows(scenario.excluded_privilege_roles),
            scenario.sheet_comments.get("Privilege-Roles", ""),
        )
        workbook.save(path)

    @staticmethod
    def _privilege_role_rows(items: list[dict[str, str]]) -> list[list[str]]:
        rows: list[list[str]] = []
        active_category = ""
        for item in items:
            category = item.get("category", "").strip()
            if category and category != active_category:
                rows.append([f"# {category}", ""])
                active_category = category
            rows.append([item.get("privilege", ""), item.get("role", "")])
        return rows

    @staticmethod
    def _write_scenario_sheet(worksheet, headers: list[str], rows: list[list[str]], comment: str) -> None:
        comment = comment.strip() or DEFAULT_SCENARIO_COMMENT
        if comment and not comment.startswith("#"):
            comment = f"# {comment}"
        header_row = 2
        if comment:
            worksheet.cell(1, 1, comment)
            worksheet.cell(1, 1).font = Font(italic=True, color="667085")
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(header_row, column, header)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left")
        for row_index, values in enumerate(rows, start=header_row + 1):
            for column, value in enumerate(values, start=1):
                worksheet.cell(row_index, column, value).alignment = Alignment(horizontal="left")
                if str(values[0]).strip().startswith("#"):
                    worksheet.cell(row_index, column).font = Font(
                        bold=True, italic=True, color="1F4E78"
                    )
                    worksheet.cell(row_index, column).fill = PatternFill(
                        "solid", fgColor="D9EAF7"
                    )
        widths = {"Privilege": 54, "Role": 62}
        for column, header in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(column)].width = widths[header]
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, header_row + len(rows))}"
        worksheet.sheet_view.showGridLines = False

    def load_saved(self) -> None:
        self.scenarios = []
        if not self.storage_dir.exists():
            return
        for path in self.storage_dir.glob("*.xlsx"):
            try:
                self.scenarios.append(self._load_scenario_workbook(path))
            except (OSError, ValueError, TypeError, KeyError):
                continue

    @staticmethod
    def _load_scenario_workbook(path: Path) -> Scenario:
        workbook = load_workbook(path, read_only=False, data_only=True)
        missing = [name for name in SCENARIO_SHEETS if name not in workbook.sheetnames]
        extra = [name for name in workbook.sheetnames if name not in SCENARIO_SHEETS]
        if missing or extra:
            details = []
            if missing:
                details.append(f"missing: {', '.join(missing)}")
            if extra:
                details.append(f"unexpected: {', '.join(extra)}")
            raise ValueError(f"Scenario workbook must contain exactly three sheets ({'; '.join(details)})")

        def read_sheet(sheet_name: str, required_headers: list[str]) -> tuple[list[dict[str, str]], str]:
            worksheet = workbook[sheet_name]
            first_value = str(worksheet.cell(1, 1).value or "").strip()
            comment = first_value if first_value.startswith("#") else ""
            header_row = 2
            headers = {
                str(worksheet.cell(header_row, column).value or "").strip().casefold(): column
                for column in range(1, worksheet.max_column + 1)
            }
            absent = [header for header in required_headers if header.casefold() not in headers]
            if absent:
                raise ValueError(f"{sheet_name} is missing column(s): {', '.join(absent)}")
            rows: list[dict[str, str]] = []
            for row_index in range(header_row + 1, worksheet.max_row + 1):
                first_value = str(worksheet.cell(row_index, 1).value or "").strip()
                if first_value.startswith("#"):
                    continue
                item = {
                    header: str(
                        worksheet.cell(row_index, headers[header.casefold()]).value or ""
                    ).strip()
                    for header in required_headers
                }
                if any(item.values()):
                    rows.append(item)
            return rows, comment

        role_rows, role_comment = read_sheet("Roles", ["Role"])
        privilege_rows, privilege_comment = read_sheet("Privileges", ["Privilege"])
        pair_sheet = workbook["Privilege-Roles"]
        pair_first_value = str(pair_sheet.cell(1, 1).value or "").strip()
        pair_comment = pair_first_value if pair_first_value.startswith("#") else ""
        pair_headers = {
            str(pair_sheet.cell(2, column).value or "").strip().casefold(): column
            for column in range(1, pair_sheet.max_column + 1)
        }
        if "privilege" not in pair_headers or "role" not in pair_headers:
            raise ValueError("Privilege-Roles is missing Privilege or Role")
        pair_rows: list[dict[str, str]] = []
        category = ""
        for row_index in range(3, pair_sheet.max_row + 1):
            privilege = str(
                pair_sheet.cell(row_index, pair_headers["privilege"]).value or ""
            ).strip()
            if privilege.startswith("#"):
                category = privilege[1:].strip()
                continue
            role = str(pair_sheet.cell(row_index, pair_headers["role"]).value or "").strip()
            if privilege and role:
                item = {"privilege": privilege, "role": role}
                if category:
                    item["category"] = category
                pair_rows.append(item)
        metadata = {
            str(workbook["Roles"].cell(row, 4).value or "").strip(): str(
                workbook["Roles"].cell(row, 5).value or ""
            ).strip()
            for row in range(1, 5)
        }
        missing_metadata = [
            label for label in ("Scenario", "SKU", "Service", "Created at") if not metadata.get(label)
        ]
        if missing_metadata:
            workbook.close()
            raise ValueError(f"Roles sheet is missing metadata: {', '.join(missing_metadata)}")
        created_at = metadata.get("Created at") or datetime.now().isoformat(timespec="seconds")
        scenario = Scenario(
            name=path.stem,
            sku=metadata.get("SKU", ""),
            service=metadata.get("Service", ""),
            excluded_roles=sorted({item["Role"] for item in role_rows if item["Role"]}),
            excluded_privileges=sorted(
                {item["Privilege"] for item in privilege_rows if item["Privilege"]}
            ),
            excluded_privilege_roles=pair_rows,
            sheet_comments={
                name: comment
                for name, comment in (
                    ("Roles", role_comment),
                    ("Privileges", privilege_comment),
                    ("Privilege-Roles", pair_comment),
                )
                if comment
            },
            created_at=created_at,
        )
        workbook.close()
        return scenario


class ScenarioPicker(tk.Toplevel):
    """Modal picker that prevents users having to retype an existing name."""
    def __init__(
        self, parent: tk.Misc, scenarios: list[Scenario], *, x_root: int | None = None,
        y_root: int | None = None
    ) -> None:
        super().__init__(parent)
        self.title("Choose scenario")
        self.resizable(False, False)
        self.result: str | None = None
        self.scenarios = scenarios
        self.choice_var = tk.StringVar(value="Create new scenario…")
        self.name_var = tk.StringVar()
        body = ttk.Frame(self, padding=12)
        body.pack(fill="both", expand=True)
        ttk.Label(body, text="Add the selected exclusions to").grid(row=0, column=0, sticky="w")
        self.choice = ttk.Combobox(body, textvariable=self.choice_var, state="readonly", width=42)
        self.choice["values"] = [item.name for item in scenarios] + ["Create new scenario…"]
        self.choice.grid(row=1, column=0, sticky="ew", pady=(2, 10))
        ttk.Label(body, text="New scenario name (only when creating a new scenario)").grid(row=2, column=0, sticky="w")
        name_entry = ttk.Entry(body, textvariable=self.name_var, width=45)
        name_entry.grid(row=3, column=0, sticky="ew", pady=(2, 12))
        name_entry.bind("<KeyRelease>", lambda _event: self.choice_var.set("Create new scenario…"))
        buttons = ttk.Frame(body)
        buttons.grid(row=4, column=0, sticky="e")
        ttk.Button(buttons, text="Cancel", command=self.destroy).pack(side="right")
        ttk.Button(buttons, text="Use scenario", command=self._accept).pack(side="right", padx=(0, 6))
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.transient(parent)
        self.update_idletasks()
        if x_root is not None and y_root is not None:
            width, height = self.winfo_reqwidth(), self.winfo_reqheight()
            x = min(x_root + 8, self.winfo_screenwidth() - width - 12)
            y = min(y_root + 8, self.winfo_screenheight() - height - 36)
            self.geometry(f"+{max(0, x)}+{max(0, y)}")
        self.grab_set()
        name_entry.focus_set()

    def _accept(self) -> None:
        choice = self.choice_var.get()
        if choice == "Create new scenario…":
            choice = self.name_var.get().strip()
            if not choice:
                messagebox.showwarning("Scenario name required", "Enter a name for the new scenario.", parent=self)
                return
        self.result = choice
        self.destroy()


class ScenarioRolePicker(tk.Toplevel):
    """Choose one or more roles that contain a selected privilege."""

    def __init__(
        self,
        parent: tk.Misc,
        privilege: str,
        roles: list[tuple[str, int]],
        *,
        x_root: int | None = None,
        y_root: int | None = None,
    ) -> None:
        super().__init__(parent)
        self.title("Choose role(s)")
        self.geometry("760x460")
        self.minsize(560, 320)
        self.result: list[str] = []
        self.roles = roles
        self.search_var = tk.StringVar()

        body = ttk.Frame(self, padding=12)
        body.pack(fill="both", expand=True)
        ttk.Label(
            body,
            text=f"Exclude {privilege} only from the selected role(s).",
            wraplength=720,
        ).pack(anchor="w")
        ttk.Label(body, text="Search roles").pack(anchor="w", pady=(10, 2))
        search = ttk.Entry(body, textvariable=self.search_var)
        search.pack(fill="x")
        search.bind("<KeyRelease>", lambda _event: self._fill())

        table = ttk.Frame(body)
        table.pack(fill="both", expand=True, pady=(8, 10))
        self.tree = ttk.Treeview(
            table,
            columns=("role", "users"),
            show="headings",
            selectmode="extended",
        )
        self.tree.heading("role", text="Role", anchor="w")
        self.tree.heading("users", text="Users", anchor="e")
        self.tree.column("role", width=610, anchor="w", stretch=True)
        self.tree.column("users", width=90, anchor="e", stretch=False)
        scrollbar = ttk.Scrollbar(table, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", lambda _event: self._accept())

        actions = ttk.Frame(body)
        actions.pack(fill="x")
        ttk.Label(actions, text="Ctrl/Shift+click to select multiple roles.").pack(side="left")
        ttk.Button(actions, text="Cancel", command=self.destroy).pack(side="right")
        ttk.Button(actions, text="Use selected role(s)", command=self._accept).pack(
            side="right", padx=(0, 6)
        )

        self._fill()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.transient(parent)
        self.update_idletasks()
        if x_root is not None and y_root is not None:
            width, height = self.winfo_width(), self.winfo_height()
            x = min(x_root + 8, self.winfo_screenwidth() - width - 12)
            y = min(y_root + 8, self.winfo_screenheight() - height - 36)
            self.geometry(f"{width}x{height}+{max(0, x)}+{max(0, y)}")
        self.grab_set()
        search.focus_set()

    def _fill(self) -> None:
        selected_roles = {
            str(self.tree.item(item_id, "values")[0]) for item_id in self.tree.selection()
        }
        self.tree.delete(*self.tree.get_children())
        query = self.search_var.get().strip().casefold()
        for index, (role, users) in enumerate(self.roles):
            if query and query not in role.casefold():
                continue
            item_id = f"role-{index}"
            self.tree.insert("", "end", iid=item_id, values=(role, f"{users:,}"))
            if role in selected_roles:
                self.tree.selection_add(item_id)

    def _accept(self) -> None:
        self.result = [str(self.tree.item(item_id, "values")[0]) for item_id in self.tree.selection()]
        if not self.result:
            messagebox.showinfo("Choose a role", "Select at least one role.", parent=self)
            return
        self.destroy()


class ScenarioEditor(tk.Toplevel):
    def __init__(self, parent: tk.Misc, scenario: Scenario, on_changed) -> None:
        super().__init__(parent)
        self.title(f"Edit scenario — {scenario.name}")
        self.geometry("980x440")
        self.minsize(720, 300)
        self.scenario, self.on_changed = scenario, on_changed
        self.privilege_role_items: dict[str, dict[str, str]] = {}
        ttk.Label(
            self,
            text="Remove exclusions that should no longer apply.",
            padding=10,
            anchor="w",
            justify="left",
        ).pack(fill="x")
        comments = "    ".join(
            f"{name}: {comment}"
            for name, comment in scenario.sheet_comments.items()
            if comment
        )
        if comments:
            ttk.Label(
                self,
                text=comments,
                padding=(10, 0, 10, 8),
                anchor="w",
                justify="left",
                wraplength=940,
            ).pack(fill="x")
        self.tree = ttk.Treeview(
            self,
            columns=("type", "category", "name", "role"),
            show="headings",
            selectmode="extended",
        )
        self.tree.heading("type", text="Type", anchor="w")
        self.tree.heading("category", text="Category", anchor="w")
        self.tree.heading("name", text="Excluded role or privilege", anchor="w")
        self.tree.heading(
            "role", text="Role for role-specific privilege exclusion", anchor="w"
        )
        self.tree.column("type", width=150, stretch=False, anchor="w")
        self.tree.column("category", width=130, stretch=False, anchor="w")
        self.tree.column("name", width=350, stretch=True, anchor="w")
        self.tree.column("role", width=330, stretch=True, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        buttons = ttk.Frame(self, padding=(10, 0, 10, 10))
        buttons.pack(fill="x")
        ttk.Button(buttons, text="Remove selected exclusions", command=self.remove_selected).pack(side="left")
        ttk.Button(buttons, text="Close", command=self.destroy).pack(side="right")
        self.fill()

    def fill(self) -> None:
        self.tree.delete(*self.tree.get_children())
        self.privilege_role_items = {}
        for value in self.scenario.excluded_roles:
            self.tree.insert("", "end", iid=f"role:{value}", values=("Role", "", value, ""))
        for value in self.scenario.excluded_privileges:
            self.tree.insert("", "end", iid=f"privilege:{value}", values=("Privilege", "", value, ""))
        for index, item in enumerate(self.scenario.excluded_privilege_roles):
            item_id = f"privilege-role:{index}"
            self.privilege_role_items[item_id] = item
            self.tree.insert(
                "",
                "end",
                iid=item_id,
                values=(
                    "Privilege from role",
                    item.get("category", ""),
                    item.get("privilege", ""),
                    item.get("role", ""),
                ),
            )

    def remove_selected(self) -> None:
        for node_id in self.tree.selection():
            if node_id in self.privilege_role_items:
                item = self.privilege_role_items[node_id]
                if item in self.scenario.excluded_privilege_roles:
                    self.scenario.excluded_privilege_roles.remove(item)
                continue
            kind, value = node_id.split(":", 1)
            target = self.scenario.excluded_roles if kind == "role" else self.scenario.excluded_privileges
            if value in target:
                target.remove(value)
        self.fill()
        self.on_changed()


class ScenarioExclusionsDialog(tk.Toplevel):
    """Read-only, scrollable view of every exclusion in a scenario."""

    def __init__(self, parent: tk.Misc, scenario: Scenario) -> None:
        super().__init__(parent)
        self.title(f"Scenario exclusions — {scenario.name}")
        self.geometry("1020x500")
        self.minsize(720, 320)
        self.transient(parent)

        ttk.Label(
            self,
            text=f"{scenario.name}: {len(scenario.excluded_roles):,} excluded role(s), "
            f"{len(scenario.excluded_privileges):,} excluded privilege(s), and "
            f"{len(scenario.excluded_privilege_roles):,} role-specific privilege exclusion(s)",
            padding=(10, 10, 10, 8),
        ).pack(anchor="w")

        table = ttk.Frame(self)
        table.pack(fill="both", expand=True, padx=10)
        tree = ttk.Treeview(
            table, columns=("type", "category", "name", "role"), show="headings"
        )
        tree.heading("type", text="Type", anchor="w")
        tree.heading("category", text="Category", anchor="w")
        tree.heading("name", text="Excluded role or privilege", anchor="w")
        tree.heading("role", text="Role for role-specific privilege exclusion", anchor="w")
        tree.column("type", width=150, anchor="w", stretch=False)
        tree.column("category", width=130, anchor="w", stretch=False)
        tree.column("name", width=360, anchor="w", stretch=True)
        tree.column("role", width=360, anchor="w", stretch=True)
        vertical = ttk.Scrollbar(table, orient="vertical", command=tree.yview)
        horizontal = ttk.Scrollbar(table, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vertical.set, xscrollcommand=horizontal.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vertical.grid(row=0, column=1, sticky="ns")
        horizontal.grid(row=1, column=0, sticky="ew")
        table.rowconfigure(0, weight=1)
        table.columnconfigure(0, weight=1)

        for value in scenario.excluded_roles:
            tree.insert("", "end", values=("Role", "", value, ""))
        for value in scenario.excluded_privileges:
            tree.insert("", "end", values=("Privilege", "", value, ""))
        for item in scenario.excluded_privilege_roles:
            tree.insert(
                "",
                "end",
                values=(
                    "Privilege from role",
                    item.get("category", ""),
                    item.get("privilege", ""),
                    item.get("role", ""),
                ),
            )

        buttons = ttk.Frame(self, padding=10)
        buttons.pack(fill="x")
        ttk.Button(buttons, text="Close", command=self.destroy).pack(side="right")
        self.bind("<Escape>", lambda _event: self.destroy())


class ScenarioReport(tk.Toplevel):
    """Independent window for comparing baseline and saved scenarios."""
    def __init__(self, parent: tk.Misc, library: ScenarioLibrary, frame: pd.DataFrame, sku: str, service: str) -> None:
        super().__init__(parent)
        self.title("Scenario analysis report")
        self.geometry("1050x620")
        self.minsize(820, 480)
        self.library, self.frame, self.sku, self.service = library, frame.copy(), sku, service
        self.available = self.library.for_sku(sku, service)
        self.selection_vars = [tk.StringVar(), tk.StringVar()]
        self.detail_scenarios: dict[str, Scenario] = {}
        self.current_scenarios: list[Scenario | None] = []
        self._details_after_id: str | None = None
        self._resizing_detail_column = False
        self._build()
        self.refresh()

    def _build(self) -> None:
        head = ttk.Frame(self, padding=10)
        head.pack(fill="x")
        ttk.Label(head, text="Scenario analysis report", font=("Segoe UI", 15, "bold")).pack(side="left")
        ttk.Label(head, text=f"{self.sku} | {self.service}").pack(side="left", padx=14)
        ttk.Button(head, text="Reload saved scenarios", command=self.reload_saved).pack(side="right")
        ttk.Button(head, text="Edit selected scenario…", command=self.edit_selected).pack(side="right", padx=6)

        chooser = ttk.LabelFrame(self, text="Compare with the unrestricted baseline", padding=8)
        chooser.pack(fill="x", padx=10, pady=(0, 8))
        for index, variable in enumerate(self.selection_vars, start=1):
            ttk.Label(chooser, text=f"Scenario {index}").grid(row=0, column=(index - 1) * 2, sticky="w")
            combo = ttk.Combobox(chooser, textvariable=variable, state="readonly", width=42)
            combo.grid(row=1, column=(index - 1) * 2, padx=(0, 12), sticky="ew")
            combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh())
            setattr(self, f"combo_{index}", combo)
        ttk.Button(chooser, text="Refresh", command=self.refresh).grid(row=1, column=4, sticky="w")
        chooser.columnconfigure(0, weight=1)
        chooser.columnconfigure(2, weight=1)

        self.summary = ttk.Treeview(self, columns=("metric", "baseline", "first", "second"), show="headings", height=5)
        for column, heading, width in (("metric", "Measure", 285), ("baseline", "Baseline", 150), ("first", "Scenario 1", 220), ("second", "Scenario 2", 220)):
            self.summary.heading(column, text=heading, anchor="w")
            self.summary.column(column, width=width, anchor="w", stretch=column == "metric")
        self.summary.pack(fill="x", padx=10)

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)
        details = ttk.Frame(notebook, padding=7)
        notebook.add(details, text="Exclusions")
        self.detail_tree = ttk.Treeview(
            details,
            columns=("scenario", "roles", "privileges", "privilege_roles"),
            show="headings",
            height=3,
        )
        for column, heading, width in (
            ("scenario", "Scenario", 170),
            ("roles", "Excluded roles", 260),
            ("privileges", "Excluded privileges", 260),
            ("privilege_roles", "Privileges excluded from roles", 310),
        ):
            self.detail_tree.heading(column, text=heading, anchor="w")
            self.detail_tree.column(column, width=width, anchor="w", stretch=True)
        self.detail_tree.pack(fill="both", expand=True)
        self.detail_tree.bind("<Double-1>", self._open_clicked_exclusions)
        self.detail_tree.bind("<Return>", lambda _event: self.open_selected_exclusions())
        self.detail_tree.bind("<Configure>", self._schedule_detail_refresh)
        self.detail_tree.bind("<ButtonPress-1>", self._start_detail_column_resize, add="+")
        self.detail_tree.bind("<ButtonRelease-1>", self._finish_detail_column_resize, add="+")
        detail_actions = ttk.Frame(details)
        detail_actions.pack(fill="x", pady=(6, 0))
        ttk.Label(
            detail_actions,
            text="Long lists are abbreviated with …. Double-click a scenario to view every exclusion.",
        ).pack(side="left")
        ttk.Button(
            detail_actions,
            text="View complete exclusions…",
            command=self.open_selected_exclusions,
        ).pack(side="right")

        users = ttk.Frame(notebook, padding=7)
        users.rowconfigure(1, weight=1)
        users.columnconfigure(0, weight=1)
        notebook.add(users, text="User impact")
        self.user_impact_var = tk.StringVar(value="Select one or two scenarios to compare affected users.")
        ttk.Label(users, textvariable=self.user_impact_var).grid(row=0, column=0, sticky="w", pady=(0, 5))
        user_columns = (
            "USER", "BASELINE", "S1_EXCLUDED", "S1_REMAINING", "S1_STATUS",
            "S2_EXCLUDED", "S2_REMAINING", "S2_STATUS",
        )
        self.user_tree = ttk.Treeview(users, columns=user_columns, show="headings")
        for column, heading, width in (
            ("USER", "User (hashed)", 255),
            ("BASELINE", "Baseline relationships", 135),
            ("S1_EXCLUDED", "Scenario 1 excluded", 125),
            ("S1_REMAINING", "Scenario 1 remaining", 130),
            ("S1_STATUS", "Scenario 1 status", 140),
            ("S2_EXCLUDED", "Scenario 2 excluded", 125),
            ("S2_REMAINING", "Scenario 2 remaining", 130),
            ("S2_STATUS", "Scenario 2 status", 140),
        ):
            self.user_tree.heading(column, text=heading, anchor="w", command=lambda c=column: self._sort_users(c))
            self.user_tree.column(column, width=width, anchor="w" if column in {"USER", "S1_STATUS", "S2_STATUS"} else "e")
        user_y = ttk.Scrollbar(users, orient="vertical", command=self.user_tree.yview)
        user_x = ttk.Scrollbar(users, orient="horizontal", command=self.user_tree.xview)
        self.user_tree.configure(yscrollcommand=user_y.set, xscrollcommand=user_x.set)
        self.user_tree.grid(row=1, column=0, sticky="nsew")
        user_y.grid(row=1, column=1, sticky="ns")
        user_x.grid(row=2, column=0, sticky="ew")
        self.user_tree.bind("<Double-1>", lambda _event: self.open_selected_user())
        user_actions = ttk.Frame(users)
        user_actions.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Label(user_actions, text="Double-click a user to inspect every baseline and scenario assignment.").pack(side="left")
        ttk.Button(user_actions, text="View user assignments…", command=self.open_selected_user).pack(side="right")
        ttk.Label(self, text="‘Estimated required licences’ is the remaining number of users with at least one non-excluded assignment. Confirm contractual licensing rules separately.").pack(anchor="w", padx=12, pady=(0, 8))

    def _selected(self, index: int) -> Scenario | None:
        name = self.selection_vars[index].get()
        return next((item for item in self.available if item.name == name), None)

    def refresh(self) -> None:
        self.available = self.library.for_sku(self.sku, self.service)
        names = [item.name for item in self.available]
        for index, variable in enumerate(self.selection_vars, start=1):
            combo = getattr(self, f"combo_{index}")
            combo["values"] = names
            if variable.get() not in names:
                variable.set("")
        scenarios = [self._selected(0), self._selected(1)]
        self.summary.heading("first", text=scenarios[0].name if scenarios[0] else "Scenario 1")
        self.summary.heading("second", text=scenarios[1].name if scenarios[1] else "Scenario 2")
        baseline = scenario_impact(self.frame)
        impacts = [scenario_impact(self.frame, item) if item else None for item in scenarios]
        self.summary.delete(*self.summary.get_children())
        metrics = (
            ("Estimated required licences (in-scope users)", "in_scope"),
            ("Users affected by exclusions", "affected"),
            ("Estimated licences reduced (users removed from scope)", "removed"),
        )
        for label, key in metrics:
            self.summary.insert("", "end", values=(label, f"{baseline[key]:,}", *(f"{item[key]:,}" if item else "—" for item in impacts)))
        self.current_scenarios = scenarios
        self._refresh_detail_rows()
        self._refresh_user_rows()

    def _refresh_user_rows(self) -> None:
        baseline = scenario_user_summary(self.frame).set_index("USER")
        comparisons = [
            scenario_user_summary(self.frame, scenario).set_index("USER") if scenario else None
            for scenario in self.current_scenarios
        ]
        current = pd.DataFrame(index=baseline.index)
        current["USER"] = current.index.astype(str)
        current["BASELINE"] = baseline["RELATIONSHIPS"]
        affected = pd.Series(False, index=current.index)
        for number, comparison in enumerate(comparisons, start=1):
            if comparison is None:
                current[f"S{number}_EXCLUDED"] = 0
                current[f"S{number}_REMAINING"] = baseline["RELATIONSHIPS"]
                current[f"S{number}_STATUS"] = "—"
            else:
                current[f"S{number}_EXCLUDED"] = comparison["EXCLUDED"]
                current[f"S{number}_REMAINING"] = comparison["REMAINING"]
                current[f"S{number}_STATUS"] = comparison["STATUS"]
                affected |= comparison["EXCLUDED"].gt(0)
        current = current[affected].reset_index(drop=True)
        self.current_user_rows = current.sort_values(
            ["S1_EXCLUDED", "S2_EXCLUDED", "USER"],
            ascending=[False, False, True],
            kind="stable",
        )
        self.user_tree.heading(
            "S1_STATUS",
            text=f"{self.current_scenarios[0].name} status" if self.current_scenarios[0] else "Scenario 1 status",
        )
        self.user_tree.heading(
            "S2_STATUS",
            text=f"{self.current_scenarios[1].name} status" if self.current_scenarios[1] else "Scenario 2 status",
        )
        self._fill_user_rows()

    def _fill_user_rows(self) -> None:
        self.user_tree.delete(*self.user_tree.get_children())
        for row in self.current_user_rows.itertuples(index=False):
            self.user_tree.insert(
                "", "end", iid=str(row.USER),
                values=(
                    row.USER, f"{row.BASELINE:,}", f"{row.S1_EXCLUDED:,}",
                    f"{row.S1_REMAINING:,}", row.S1_STATUS, f"{row.S2_EXCLUDED:,}",
                    f"{row.S2_REMAINING:,}", row.S2_STATUS,
                ),
            )
        self.user_impact_var.set(
            f"{len(self.current_user_rows):,} users are affected by at least one selected scenario."
        )

    def _sort_users(self, column: str) -> None:
        if getattr(self, "current_user_rows", pd.DataFrame()).empty:
            return
        numeric = column not in {"USER", "S1_STATUS", "S2_STATUS"}
        state = getattr(self, "_user_sort", ("S1_EXCLUDED", True))
        descending = not state[1] if state[0] == column else numeric
        self._user_sort = (column, descending)
        self.current_user_rows = self.current_user_rows.sort_values(
            [column, "USER"] if column != "USER" else ["USER"],
            ascending=[not descending, True] if column != "USER" else [not descending],
            kind="stable",
        )
        self._fill_user_rows()

    def open_selected_user(self) -> None:
        selected = self.user_tree.selection()
        if selected:
            UserAssignmentDialog(self, self.frame, str(selected[0]), self.current_scenarios)

    def _schedule_detail_refresh(self, _event: tk.Event | None = None) -> None:
        if self._details_after_id is not None:
            self.after_cancel(self._details_after_id)
        self._details_after_id = self.after_idle(self._refresh_detail_rows)

    def _start_detail_column_resize(self, event: tk.Event) -> None:
        self._resizing_detail_column = self.detail_tree.identify_region(event.x, event.y) == "separator"

    def _finish_detail_column_resize(self, _event: tk.Event) -> None:
        if self._resizing_detail_column:
            self._resizing_detail_column = False
            self._schedule_detail_refresh()

    def _abbreviate_exclusions(self, values: list[str], column: str) -> str:
        if not values:
            return "—"
        available_width = max(20, int(self.detail_tree.column(column, "width")) - 18)
        font = tkfont.nametofont("TkDefaultFont")
        complete = ", ".join(values)
        if font.measure(complete) <= available_width:
            return complete
        visible: list[str] = []
        for value in values:
            candidate = f"{', '.join((*visible, value))}, …"
            if font.measure(candidate) > available_width:
                break
            visible.append(value)
        return f"{', '.join(visible)}, …" if visible else "…"

    def _refresh_detail_rows(self) -> None:
        self._details_after_id = None
        self.detail_tree.delete(*self.detail_tree.get_children())
        self.detail_scenarios = {}
        for index, scenario in enumerate(self.current_scenarios):
            if scenario:
                item_id = f"scenario-{index}"
                self.detail_scenarios[item_id] = scenario
                self.detail_tree.insert(
                    "",
                    "end",
                    iid=item_id,
                    values=(
                        scenario.name,
                        self._abbreviate_exclusions(scenario.excluded_roles, "roles"),
                        self._abbreviate_exclusions(scenario.excluded_privileges, "privileges"),
                        self._abbreviate_exclusions(
                            [
                                (
                                    f"[{item.get('category')}] "
                                    if item.get("category")
                                    else ""
                                )
                                + f"{item.get('privilege', '')} → {item.get('role', '')}"
                                for item in scenario.excluded_privilege_roles
                            ],
                            "privilege_roles",
                        ),
                    ),
                )

    def _open_clicked_exclusions(self, event: tk.Event) -> None:
        item_id = self.detail_tree.identify_row(event.y)
        if item_id:
            self.detail_tree.selection_set(item_id)
            self.open_selected_exclusions()

    def open_selected_exclusions(self) -> None:
        selected = self.detail_tree.selection()
        if not selected:
            messagebox.showinfo(
                "Choose a scenario",
                "Select a scenario in the exclusions table first.",
                parent=self,
            )
            return
        scenario = self.detail_scenarios.get(selected[0])
        if scenario is not None:
            ScenarioExclusionsDialog(self, scenario)

    def edit_selected(self) -> None:
        scenario = self._selected(0) or self._selected(1)
        if scenario is None:
            messagebox.showinfo("Choose a scenario", "Select a scenario in either comparison list first.", parent=self)
            return
        ScenarioEditor(self, scenario, self._scenario_changed)

    def _scenario_changed(self) -> None:
        scenario = self._selected(0) or self._selected(1)
        if scenario:
            self.library.save_scenario(scenario)
        self.refresh()

    def reload_saved(self) -> None:
        self.library.load_saved()
        self.refresh()
