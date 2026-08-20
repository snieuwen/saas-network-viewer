import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scenario_analysis import Scenario, ScenarioLibrary, scenario_impact


class ScenarioImpactTests(unittest.TestCase):
    def setUp(self) -> None:
        self.frame = pd.DataFrame(
            [
                {"USER_LOGIN_HASH": "U1", "ROLE_CODE": "R", "PRIVILEGE": "P1"},
                {"USER_LOGIN_HASH": "U1", "ROLE_CODE": "R", "PRIVILEGE": "P2"},
                {"USER_LOGIN_HASH": "U1", "ROLE_CODE": "R", "PRIVILEGE": "P3"},
                {"USER_LOGIN_HASH": "U2", "ROLE_CODE": "R", "PRIVILEGE": "P3"},
                {"USER_LOGIN_HASH": "U3", "ROLE_CODE": "S", "PRIVILEGE": "P1"},
            ]
        )

    def test_privilege_role_exclusion_only_removes_exact_relationship(self) -> None:
        scenario = Scenario("Test", "SKU", "Service")
        scenario.add_privilege_roles("P1", ["R"])

        self.assertEqual(
            scenario_impact(self.frame, scenario),
            {"in_scope": 3, "removed": 0, "affected": 1},
        )

    def test_multiple_privileges_can_be_removed_from_one_role(self) -> None:
        scenario = Scenario("Test", "SKU", "Service")
        scenario.add_privilege_roles("P1", ["R"])
        scenario.add_privilege_roles("P2", ["R"])

        self.assertEqual(len(scenario.excluded_privilege_roles), 2)
        self.assertEqual(
            scenario_impact(self.frame, scenario),
            {"in_scope": 3, "removed": 0, "affected": 1},
        )

    def test_pair_exclusions_combine_with_global_exclusions(self) -> None:
        scenario = Scenario("Test", "SKU", "Service", excluded_privileges=["P3"])
        scenario.add_privilege_roles("P1", ["R"])
        scenario.add_privilege_roles("P2", ["R"])

        self.assertEqual(
            scenario_impact(self.frame, scenario),
            {"in_scope": 1, "removed": 2, "affected": 2},
        )

    def test_privilege_role_categories_use_business_order(self) -> None:
        scenario = Scenario("Test", "SKU", "Service")
        scenario.add_privilege_roles("P3", ["R"], "Projecten")
        scenario.add_privilege_roles("P1", ["R"], "Manage")
        scenario.add_privilege_roles("P2", ["R"], "Inkoop Raadplegen")

        self.assertEqual(
            [item["category"] for item in scenario.excluded_privilege_roles],
            ["Manage", "Inkoop Raadplegen", "Projecten"],
        )


class ScenarioLibraryTests(unittest.TestCase):
    def test_scenario_round_trip_uses_three_excel_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            scenario = Scenario(
                "Excel scenario",
                "SKU",
                "Service",
                excluded_roles=["R"],
                excluded_privileges=["P1"],
                excluded_privilege_roles=[
                    {"category": "Manage", "privilege": "P2", "role": "R"}
                ],
                sheet_comments={"Roles": "# This comment has no calculation meaning"},
                created_at="2026-08-14T00:00:00",
            )
            library = ScenarioLibrary()
            library.storage_dir = Path(directory)
            library.save_scenario(scenario)
            library.load_saved()

            self.assertEqual(len(library.scenarios), 1)
            loaded = library.scenarios[0]
            self.assertEqual(loaded.name, "Excel scenario")
            self.assertEqual(loaded.sku, "SKU")
            self.assertEqual(loaded.service, "Service")
            self.assertEqual(loaded.excluded_roles, ["R"])
            self.assertEqual(loaded.excluded_privileges, ["P1"])
            self.assertEqual(
                loaded.excluded_privilege_roles,
                [{"category": "Manage", "privilege": "P2", "role": "R"}],
            )
            self.assertEqual(
                loaded.sheet_comments["Roles"], "# This comment has no calculation meaning"
            )
            workbook = load_workbook(Path(directory) / "Excel scenario.xlsx", read_only=True)
            self.assertEqual(workbook.sheetnames, ["Roles", "Privileges", "Privilege-Roles"])
            self.assertEqual(workbook["Roles"]["D1"].value, "Scenario")
            self.assertEqual(workbook["Roles"]["E1"].value, "Excel scenario")
            self.assertEqual(workbook["Roles"]["D2"].value, "SKU")
            self.assertEqual(workbook["Roles"]["E2"].value, "SKU")
            self.assertEqual(workbook["Privilege-Roles"]["A3"].value, "# Manage")
            workbook.close()

    def test_json_files_are_not_imported(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            (Path(directory) / "Legacy.json").write_text("{}", encoding="utf-8")
            library = ScenarioLibrary()
            library.storage_dir = Path(directory)
            library.load_saved()

            self.assertEqual(library.scenarios, [])


if __name__ == "__main__":
    unittest.main()
